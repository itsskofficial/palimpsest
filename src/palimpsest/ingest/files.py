"""Local files: text, PDF, image, spreadsheet.

Each adapter produces the same thing — normalised text plus segments — but the shape of
a useful anchor differs by format, and matching it to the format is most of the value:

- **PDF** anchors to a page number, because that is how a person refers to a PDF.
- **Tabular** anchors to a cell range, because a claim drawn from a spreadsheet is
  drawn from specific cells and "row 47" is checkable in a way "the sheet" is not.
- **Image** anchors to the whole image, since a photographed whiteboard has no
  sub-structure worth pretending about.
- **Text** anchors to headings when there are any, offsets when there are not.

Tabular deserves a note. A spreadsheet's claims are rows and aggregates, not prose, so
the adapter renders it as a compact table and lets the extractor read *facts* out of
it. Handing the model raw CSV produces claims like "the file has a column named
price", which is true and useless.
"""

from __future__ import annotations

import contextlib
import csv
import io
import logging
import mimetypes
from pathlib import Path

from palimpsest.ingest import make_source
from palimpsest.types import Source

__all__ = ["from_image", "from_pdf", "from_tabular", "from_text"]

log = logging.getLogger("palimpsest.ingest.files")

#: Rows rendered into the text handed to the extractor. Beyond this a summary is
#: rendered instead — a 50,000-row sheet is a database, not a note.
MAX_ROWS = 400


def from_text(spec: str, title: str | None = None) -> Source:
    """A literal string (`text:...`), or a local text/markdown file."""
    from palimpsest.ingest.web import segments_from_markdown

    if spec.startswith("text:"):
        body = spec[5:].strip()
        return make_source("text", title or "Note", body,
                           segments=segments_from_markdown(body), extractor="inline")

    path = Path(spec)
    if not path.exists():
        # A bare string with no matching file is a note, not a typo — capturing a
        # thought by typing it is the lowest-friction path into the system.
        return make_source("text", title or "Note", spec,
                           segments=segments_from_markdown(spec), extractor="inline")

    body = path.read_text(encoding="utf-8", errors="replace")
    return make_source("text", title or path.stem, body,
                       segments=segments_from_markdown(body),
                       extractor="file", path=str(path.resolve()))


def from_pdf(spec: str) -> Source:
    """Extract text per page, anchoring each page so citations say `p. 14`."""
    try:
        from pypdf import PdfReader
    except ImportError as e:  # pragma: no cover - optional extra
        raise ImportError(
            "PDF ingestion needs pypdf: pip install 'palimpsest[pdf]'"
        ) from e

    if spec.lower().startswith(("http://", "https://")):
        import urllib.request

        with urllib.request.urlopen(spec, timeout=60) as resp:
            data = resp.read()
        reader = PdfReader(io.BytesIO(data))
        title, url, path = Path(spec.split("?")[0]).stem, spec, None
    else:
        path = str(Path(spec).resolve())
        reader = PdfReader(path)
        title, url = Path(spec).stem, None

    meta_title = ""
    with contextlib.suppress(Exception):  # pragma: no cover - metadata is often broken
        meta_title = str(getattr(reader.metadata, "title", None) or "")

    parts: list[str] = []
    segments: list[dict] = []
    cursor = 0
    for index, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:  # pragma: no cover - a broken page should not lose the file
            text = ""
        if not text.strip():
            continue
        chunk = text.strip() + "\n\n"
        parts.append(chunk)
        segments.append({"start": cursor, "end": cursor + len(chunk), "kind": "page",
                         "locator": f"p. {index}", "url": url})
        cursor += len(chunk)

    body = "".join(parts).strip()
    if not body:
        raise RuntimeError(
            f"{spec}: no extractable text. This is probably a scanned PDF — "
            "run it through OCR, or ingest the pages as images."
        )
    return make_source("pdf", str(meta_title).strip() or title, body, url=url,
                       segments=segments, pages=len(reader.pages), path=path)


def from_image(spec: str, model=None) -> Source:
    """Read an image with the model: a whiteboard, a slide, a photographed page.

    This is the one adapter that cannot work without a model, and it says so rather
    than returning an empty source — a silent empty source would produce a run that
    "succeeded" and added nothing.
    """
    if model is None:
        raise RuntimeError(
            "reading an image needs a model. Set ANTHROPIC_API_KEY, or pass "
            "--kind text and describe the image yourself."
        )
    path = Path(spec)
    data = path.read_bytes()
    media_type = mimetypes.guess_type(str(path))[0] or "image/png"

    text = model.describe_image(
        data, media_type,
        "Transcribe this image completely and faithfully.\n"
        "- Reproduce all text verbatim, preserving structure (headings, lists, tables).\n"
        "- Describe diagrams and figures in enough detail that their content is usable "
        "without the image.\n"
        "- Do not summarise, interpret, or add anything that is not present.\n"
        "- If part of it is illegible, say so explicitly rather than guessing.",
    )
    return make_source("image", path.stem, text.strip(),
                       segments=[{"start": 0, "end": len(text), "kind": "region",
                                  "locator": path.name, "url": None}],
                       path=str(path.resolve()), media_type=media_type)


def _render_table(rows: list[list[str]], sheet: str, url: str | None
                  ) -> tuple[str, list[dict]]:
    """Render rows as a readable table, anchoring each row to its cell range."""
    if not rows:
        return "", []
    header = [str(c) for c in rows[0]]
    ncols = len(header)

    def col_letter(i: int) -> str:
        letters = ""
        i += 1
        while i:
            i, rem = divmod(i - 1, 26)
            letters = chr(65 + rem) + letters
        return letters

    span = f"A–{col_letter(max(0, ncols - 1))}"
    parts = [f"# {sheet}\n\n| " + " | ".join(header) + " |\n"
             + "|" + "|".join(["---"] * ncols) + "|\n"]
    segments: list[dict] = []
    cursor = len(parts[0])

    body_rows = rows[1:MAX_ROWS + 1]
    for index, row in enumerate(body_rows, start=2):
        cells = [str(c) if c is not None else "" for c in row][:ncols]
        cells += [""] * (ncols - len(cells))
        line = "| " + " | ".join(cells) + " |\n"
        parts.append(line)
        segments.append({"start": cursor, "end": cursor + len(line), "kind": "cell",
                         "locator": f"{sheet}!{span}{index}", "url": url})
        cursor += len(line)

    if len(rows) - 1 > MAX_ROWS:
        note = (f"\n_({len(rows) - 1 - MAX_ROWS} further rows not shown; "
                f"{len(rows) - 1} data rows in total.)_\n")
        parts.append(note)
    return "".join(parts), segments


def from_tabular(spec: str) -> Source:
    """CSV, TSV or Excel, rendered as a table the extractor can read facts from."""
    path = Path(spec)
    suffix = path.suffix.lower()

    if suffix in (".csv", ".tsv"):
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
            rows = [row for row in csv.reader(fh, delimiter=delimiter)]
        text, csv_segments = _render_table(rows, path.stem, None)
        return make_source("tabular", path.stem, text, segments=csv_segments,
                           rows=max(0, len(rows) - 1), path=str(path.resolve()))

    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - optional extra
        raise ImportError(
            "Excel ingestion needs openpyxl: pip install 'palimpsest[tabular]'"
        ) from e

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    chunks: list[str] = []
    segments: list[dict] = []
    cursor = 0
    total_rows = 0
    for sheet in workbook.worksheets:
        rows = [[c for c in row] for row in sheet.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        if not rows:
            continue
        text, sheet_segments = _render_table(rows, sheet.title, None)
        for seg in sheet_segments:
            seg["start"] += cursor
            seg["end"] += cursor
        chunks.append(text + "\n")
        segments.extend(sheet_segments)
        cursor += len(text) + 1
        total_rows += max(0, len(rows) - 1)
    workbook.close()

    return make_source("tabular", path.stem, "".join(chunks).strip(), segments=segments,
                       rows=total_rows, sheets=len(chunks), path=str(path.resolve()))
